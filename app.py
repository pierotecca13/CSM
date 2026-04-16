"""
CSM Dashboard — Backend + Frontend en un solo servidor
Incluye: JIRA Cloud, MySQL, exportación Excel/PDF/PPTX
Correr: python app.py
"""
import os, math, socket, json, io
from datetime import datetime, timezone, timedelta
from collections import defaultdict

import requests
from requests.auth import HTTPBasicAuth
from flask import Flask, jsonify, request, send_from_directory, send_file
from flask_cors import CORS

app = Flask(__name__, static_folder="static", static_url_path="/static")
CORS(app)

DEFAULT_URL   = os.getenv("JIRA_URL",   "")
DEFAULT_EMAIL = os.getenv("JIRA_EMAIL", "")
DEFAULT_TOKEN = os.getenv("JIRA_TOKEN", "")

# ── Cache ─────────────────────────────────────────────────────────
CACHE_FILE = os.path.join(os.path.dirname(__file__), "cache.json")

def save_cache(payload: dict):
    try:
        with open(CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False)
        print(f"[CACHE] Guardado en {CACHE_FILE}")
    except Exception as e:
        print(f"[CACHE] Error al guardar: {e}")

def load_cache() -> dict | None:
    try:
        if not os.path.exists(CACHE_FILE):
            return None
        with open(CACHE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"[CACHE] Error al leer: {e}")
        return None

# Campos candidatos para detectar el cliente/organización
# Orden de prioridad: primero los confirmados de esta instancia
ACCOUNT_CANDIDATES = (
    "organizations",
    "customfield_11302",  # ✓ "Organizaciones" — confirmado en esta instancia
    "customfield_11476",  # "Organización vinculada"
    "customfield_11329",  # "Account"
    "customfield_10050", "customfield_10051",
    "customfield_10072", "customfield_10014",
)

OPEN_STATUSES = {
    "open", "in progress", "waiting for customer",
    "en progreso", "escalado", "pendiente", "reopened",
    "to do", "in review", "waiting for support",
    "esperando al cliente", "en curso",
}
CLOSED_STATUSES = {
    "done", "resolved", "closed", "cerrado", "resuelto",
    "completado", "finalizado", "cancelado",
}

# ── Campos SLA confirmados en esta instancia ──────────────────────
# Orden de prioridad para auto-detección
SLA_FIELD_CANDIDATES = [
    "customfield_11418",  # ✓ Tiempo absoluto de tickets  ← principal
    "customfield_11324",  # ✓ Tiempo hasta resolución
    "customfield_11325",  # ✓ Tiempo hasta primera respuesta
    "customfield_11367",  # ✓ Time to resolution
    "customfield_11804",  # ✓ Tiempo contractual
    "customfield_11469",  # ✓ Tiempo hasta resolución máxima
    "customfield_10020",  # JSM default
]

# Campos extra a incluir siempre en el fetch
EXTRA_FIELDS = [
    "customfield_11418",  # Tiempo absoluto de tickets
    "customfield_11324",  # Tiempo hasta resolución
    "customfield_11325",  # Tiempo hasta primera respuesta
    "customfield_11502",  # Motivo del ticket
    "customfield_11489",  # Motivo de ticket
    "customfield_11340",  # Motivo de Ticket HD
    "customfield_11341",  # Motivo de Tickets
    "customfield_11387",  # Centro de costos
    "customfield_11302",  # Organizaciones
    "customfield_11476",  # Organización vinculada
    "customfield_11329",  # Account
]

# ── Servir el dashboard ───────────────────────────────────────────
@app.route("/")
def index():
    return send_from_directory("static", "index.html")

# ── JIRA helpers ─────────────────────────────────────────────────
def jira_get(url, email, token, endpoint, params=None):
    r = requests.get(
        f"{url.rstrip('/')}/rest/api/3/{endpoint}",
        auth=HTTPBasicAuth(email, token),
        headers={"Accept": "application/json"},
        params=params, timeout=20,
    )
    r.raise_for_status()
    return r.json()

def jira_search(url, email, token, jql, fields):
    """Usa el nuevo endpoint /rest/api/3/search/jql con nextPageToken."""
    results = []
    next_page_token = None
    page = 0
    while True:
        params = {
            "jql": jql,
            "fields": ",".join(fields),
            "maxResults": 100,
        }
        if next_page_token:
            params["nextPageToken"] = next_page_token

        data = jira_get(url, email, token, "search/jql", params)
        issues = data.get("issues", [])
        results.extend(issues)
        page += 1
        print(f"[JIRA] Página {page} → {len(issues)} issues (acumulado={len(results)})")

        next_page_token = data.get("nextPageToken")
        if not issues or len(issues) < 100 or not next_page_token:
            break
    return results

# ── Detección cliente ─────────────────────────────────────────────
def detect_client_field(fields_sample):
    """Analiza una muestra de issues para detectar qué campo tiene info de cliente."""
    for f in fields_sample:
        # Probar organizations (JSM)
        orgs = f.get("organizations")
        if orgs and isinstance(orgs, list) and orgs:
            return "organizations"
        # Probar customfields
        for c in ACCOUNT_CANDIDATES[1:]:
            v = f.get(c)
            if v and isinstance(v, (dict, list)):
                if isinstance(v, list) and v:
                    return c
                if isinstance(v, dict) and (v.get("name") or v.get("displayName") or v.get("value")):
                    return c
    return None

def get_client_name(f, field_key):
    if not field_key:
        return "Sin cliente"
    val = f.get(field_key)
    if not val:
        return "Sin cliente"
    if isinstance(val, list):
        names = [x.get("name") or x.get("displayName","") for x in val if x]
        return names[0] if names else "Sin cliente"
    if isinstance(val, dict):
        return (val.get("name") or val.get("displayName") or
                val.get("value") or "Sin cliente")
    return str(val) if val else "Sin cliente"

# ── SLA ───────────────────────────────────────────────────────────
def parse_sla(f, sla_field):
    if not sla_field:
        return None
    sla = f.get(sla_field)
    if not sla or not isinstance(sla, dict):
        return None
    ongoing   = sla.get("ongoingCycle")
    completed = sla.get("completedCycles", [])
    if ongoing:
        return ongoing.get("breached", False)
    if completed:
        return completed[-1].get("breached", False)
    return None

def try_detect_sla_field(sample_fields, candidates):
    """Detecta qué campo SLA tiene datos."""
    for f in sample_fields:
        for c in candidates:
            val = f.get(c)
            if val and isinstance(val, dict) and (
                val.get("ongoingCycle") or val.get("completedCycles")
            ):
                return c
    return None

def resolution_hours(f):
    c, r = f.get("created"), f.get("resolutiondate")
    if not c or not r:
        return None
    try:
        def parse_dt(s):
            s = s[:23].replace("T", " ")
            return datetime.strptime(s, "%Y-%m-%d %H:%M:%S.%f").replace(tzinfo=timezone.utc)
        return round((parse_dt(r) - parse_dt(c)).total_seconds() / 3600, 1)
    except Exception:
        return None

def days_open(f):
    c = f.get("created")
    if not c:
        return 0
    try:
        s = c[:23].replace("T", " ")
        dt = datetime.strptime(s, "%Y-%m-%d %H:%M:%S.%f").replace(tzinfo=timezone.utc)
        return (datetime.now(timezone.utc) - dt).days
    except Exception:
        return 0

def calc_health(sla_pct, mttr, aging_7, aging_14, reopen_rate):
    speed  = max(0, 100 - mttr * 4)
    aging  = max(0, 100 - aging_7 * 10 - aging_14 * 20)
    reopen = max(0, 100 - reopen_rate * 3)
    return max(0, min(100, round(sla_pct * 0.35 + speed * 0.25 + aging * 0.25 + reopen * 0.15)))

def get_quarter(date_str):
    try:
        month = int(date_str[5:7])
        year  = date_str[:4]
        q = (month - 1) // 3 + 1
        return f"Q{q} {year}"
    except Exception:
        return "?"

def process_issues(issues, sla_field, client_filter=None, days=90):
    # client_filter: set de nombres de clientes a incluir (None = todos)
    sample = [i["fields"] for i in issues[:20]]
    detected = detect_client_field(sample)
    print(f"[JIRA] Campo cliente detectado: {detected}")

    # Auto-detectar SLA si no se especificó uno válido
    if not sla_field or sla_field == "auto":
        sla_field = try_detect_sla_field(sample, SLA_FIELD_CANDIDATES)
        print(f"[JIRA] Campo SLA detectado: {sla_field}")
    else:
        # Verificar que el campo especificado tiene datos
        found = try_detect_sla_field(sample, [sla_field])
        if not found:
            # Intentar detectar automáticamente
            alt = try_detect_sla_field(sample, SLA_FIELD_CANDIDATES)
            if alt:
                print(f"[JIRA] SLA field '{sla_field}' sin datos → usando '{alt}'")
                sla_field = alt

    data = defaultdict(lambda: {
        "total": 0, "open": 0, "closed": 0,
        "sla_good": 0, "sla_bad": 0,
        "res_hours": [], "aging_7": 0, "aging_14": 0,
        "oldest": 0, "reopen": 0,
        "by_priority": defaultdict(int),
        "by_type": defaultdict(int),
        "by_month": defaultdict(int),
        "by_week": defaultdict(lambda: {"new": 0, "closed": 0}),
        "by_quarter": defaultdict(int),
        "by_motivo": defaultdict(int),
        "by_centro": defaultdict(int),
        "by_day": defaultdict(int),
        # Primera respuesta
        "first_response_good": 0, "first_response_bad": 0,
        "first_response_hours": [],
        # Críticos solamente
        "critical_total": 0,
        "critical_sla_good": 0, "critical_sla_bad": 0,
        "critical_by_priority": defaultdict(int),
        "critical_res_hours": [],
        # Últimos 30 días
        "last30_total": 0, "last30_critical": 0,
        "last30_resolved": 0,
        "last30_sla_good": 0, "last30_sla_bad": 0,
        "last30_critical_res_hours": [],
    })

    CRITICAL_PRIORITIES = {"critical", "highest", "urgente", "crítico", "crítica"}
    MOTIVO_FIELDS = [
        "customfield_11502",
        "customfield_11489",
        "customfield_11340",
        "customfield_11341",
    ]
    CENTRO_FIELD   = "customfield_11387"
    FIRST_RESP_SLA = "customfield_11325"
    TZ_ARGENTINA   = timezone(timedelta(hours=-3))
    now_utc        = datetime.now(TZ_ARGENTINA)

    for issue in issues:
        f = issue["fields"]

        # ── Detección dual: organization O account ──────────
        client = "Sin cliente"
        # 1) Intentar organizations (JSM)
        orgs = f.get("organizations")
        if orgs and isinstance(orgs, list) and orgs:
            client = orgs[0].get("name") or orgs[0].get("displayName") or "Sin cliente"
        # 2) Si no, intentar customfield_11302 (Organizaciones)
        if client == "Sin cliente":
            cf_org = f.get("customfield_11302")
            if cf_org and isinstance(cf_org, list) and cf_org:
                client = cf_org[0].get("name") or cf_org[0].get("displayName") or "Sin cliente"
        # 3) Si no, intentar Account / Organización vinculada
        if client == "Sin cliente":
            for acc_field in ("customfield_11329", "customfield_11476"):
                val = f.get(acc_field)
                if val:
                    if isinstance(val, dict):
                        name = val.get("name") or val.get("displayName") or ""
                    elif isinstance(val, list) and val:
                        name = (val[0].get("name") or val[0].get("displayName") or "") if isinstance(val[0], dict) else str(val[0])
                    else:
                        name = str(val)
                    if name:
                        client = name
                        break

        # Si se definió filtro de clientes, saltar los que no están
        if client_filter and client not in client_filter and client != "Sin cliente":
            if not any(f_name.lower() in client.lower() or client.lower() in f_name.lower()
                       for f_name in client_filter):
                continue

        # ── Fechas ───────────────────────────────────────────
        created = f.get("created", "")
        is_last30 = False
        if created:
            try:
                s = created[:23].replace("T", " ")
                created_dt = datetime.strptime(s, "%Y-%m-%d %H:%M:%S.%f").replace(tzinfo=timezone.utc)
                created_dt_aware = created_dt.astimezone(TZ_ARGENTINA)
                is_last30 = (now_utc - created_dt_aware).days <= 30
            except Exception:
                pass

        d = data[client]
        d["total"] += 1

        status = (f.get("status") or {}).get("name", "").lower()
        is_open = status in OPEN_STATUSES or status not in CLOSED_STATUSES
        if is_open:
            d["open"] += 1
            age = days_open(f)
            if age > 7:  d["aging_7"]  += 1
            if age > 14: d["aging_14"] += 1
            d["oldest"] = max(d["oldest"], age)
        else:
            d["closed"] += 1

        # ── SLA ──────────────────────────────────────────────
        breached = parse_sla(f, sla_field)
        if breached is True:   d["sla_bad"]  += 1
        elif breached is False: d["sla_good"] += 1

        if is_last30:
            d["last30_total"] += 1
            if breached is True:   d["last30_sla_bad"]  += 1
            elif breached is False: d["last30_sla_good"] += 1
            if not is_open:
                d["last30_resolved"] += 1

        # ── Primera respuesta ────────────────────────────────
        fr_breached = parse_sla(f, FIRST_RESP_SLA)
        if fr_breached is True:   d["first_response_bad"]  += 1
        elif fr_breached is False: d["first_response_good"] += 1
        fr_sla = f.get(FIRST_RESP_SLA)
        if fr_sla and isinstance(fr_sla, dict):
            cycle = fr_sla.get("completedCycles", [{}])
            if cycle:
                elapsed = cycle[-1].get("elapsedTime", {})
                if elapsed:
                    millis = elapsed.get("millis", 0)
                    if millis:
                        d["first_response_hours"].append(round(millis / 3600000, 1))

        h = resolution_hours(f)
        if h is not None:
            d["res_hours"].append(h)

        if "reopen" in status:
            d["reopen"] += 1

        d["by_priority"][(f.get("priority") or {}).get("name", "Sin prioridad")] += 1
        d["by_type"][(f.get("issuetype") or {}).get("name", "Sin tipo")] += 1

        # ── Críticos ─────────────────────────────────────────
        prio_name = (f.get("priority") or {}).get("name", "")
        is_critical = prio_name.lower() in CRITICAL_PRIORITIES
        if is_critical:
            d["critical_total"] += 1
            d["critical_by_priority"][prio_name] += 1
            breached_c = parse_sla(f, sla_field)
            if breached_c is True:   d["critical_sla_bad"]  += 1
            elif breached_c is False: d["critical_sla_good"] += 1
            h_c = resolution_hours(f)
            if h_c is not None:
                d["critical_res_hours"].append(h_c)
            if is_last30:
                d["last30_critical"] += 1
                if h_c is not None:
                    d["last30_critical_res_hours"].append(h_c)

        # ── Fechas / agrupaciones ────────────────────────────
        if created:
            d["by_month"][created[:7]] += 1
            d["by_quarter"][get_quarter(created)] += 1
            try:
                dt   = datetime.strptime(created[:10], "%Y-%m-%d")
                week = f"S{dt.isocalendar()[1]:02d} {dt.strftime('%b')}"
                d["by_week"][week]["new"] += 1
                if not is_open:
                    d["by_week"][week]["closed"] += 1
            except Exception:
                pass
            d["by_day"][created[:10]] += 1

        # ── Motivo ───────────────────────────────────────────
        for mc in MOTIVO_FIELDS:
            mv = f.get(mc)
            if mv:
                if isinstance(mv, dict):
                    motivo = mv.get("value") or mv.get("name") or ""
                elif isinstance(mv, list) and mv:
                    motivo = (mv[0].get("value") or mv[0].get("name") or "") if isinstance(mv[0], dict) else str(mv[0])
                else:
                    motivo = str(mv)
                if motivo:
                    d["by_motivo"][motivo] += 1
                    break

        # ── Centro de costos ─────────────────────────────────
        cv = f.get(CENTRO_FIELD)
        if cv:
            if isinstance(cv, dict):
                centro = cv.get("value") or cv.get("name") or ""
            elif isinstance(cv, list) and cv:
                centro = (cv[0].get("value") or cv[0].get("name") or "") if isinstance(cv[0], dict) else str(cv[0])
            else:
                centro = str(cv)
            # Excluir entradas de Trazabilidad Argentina
            if centro and "trazabilidad argentina" not in centro.lower():
                d["by_centro"][centro] += 1

    result = {}
    for client, d in data.items():
        hours       = d["res_hours"]
        mttr        = round(sum(hours) / len(hours), 1) if hours else 0
        sla_total   = d["sla_good"] + d["sla_bad"]
        sla_pct     = round(d["sla_good"] / sla_total * 100, 1) if sla_total > 0 else 0
        reopen_rate = round(d["reopen"] / d["total"] * 100, 1)  if d["total"]  > 0 else 0
        bug_rate    = round(d["by_type"].get("Bug", 0) / d["total"] * 100, 1) if d["total"] > 0 else 0
        weekly      = [{"w": w, "new": v["new"], "closed": v["closed"]}
                       for w, v in sorted(d["by_week"].items())[-8:]]

        # Distribución de Tiempo hasta Resolución por buckets
        buckets = {"<4h":0,"4-8h":0,"8-24h":0,"1-3d":0,"3-7d":0,">7d":0}
        for h in hours:
            if   h < 4:    buckets["<4h"]  += 1
            elif h < 8:    buckets["4-8h"] += 1
            elif h < 24:   buckets["8-24h"]+= 1
            elif h < 72:   buckets["1-3d"] += 1
            elif h < 168:  buckets["3-7d"] += 1
            else:          buckets[">7d"]  += 1

        # Percentiles de resolución
        sorted_h = sorted(hours)
        p50 = sorted_h[len(sorted_h)//2]         if sorted_h else 0
        p90 = sorted_h[int(len(sorted_h)*0.9)]   if sorted_h else 0
        p95 = sorted_h[int(len(sorted_h)*0.95)]  if sorted_h else 0

        result[client] = {
            "name":             client,
            "total":            d["total"],
            "open":             d["open"],
            "closed":           d["closed"],
            "sla_good":         d["sla_good"],
            "sla_bad":          d["sla_bad"],
            "sla_pct":          sla_pct,
            "mttr":             mttr,
            "aging_7":          d["aging_7"],
            "aging_14":         d["aging_14"],
            "oldest_open_days": d["oldest"],
            "reopen_rate":      reopen_rate,
            "bug_rate":         bug_rate,
            "by_priority":      dict(d["by_priority"]),
            "by_type":          dict(d["by_type"]),
            "by_month":         dict(sorted(d["by_month"].items())),
            "by_quarter":       dict(sorted(d["by_quarter"].items())),
            "by_motivo":        dict(sorted(d["by_motivo"].items(), key=lambda x: -x[1])[:10]),
            "by_centro":        dict(sorted(d["by_centro"].items(), key=lambda x: -x[1])),
            "by_day":           dict(sorted(d["by_day"].items())[-60:]),
            "weekly":           weekly,
            # Tiempo hasta resolución
            "res_buckets":      buckets,
            "res_p50":          round(p50, 1),
            "res_p90":          round(p90, 1),
            "res_p95":          round(p95, 1),
            "res_count":        len(hours),
            "health":           calc_health(sla_pct, mttr, d["aging_7"], d["aging_14"], reopen_rate),
            "sla_field_used":   sla_field or "ninguno",
            "client_field_used": detected or "ninguno",
            # Primera respuesta
            "first_response_good": d["first_response_good"],
            "first_response_bad":  d["first_response_bad"],
            "first_response_pct":  round(d["first_response_good"] /
                                   (d["first_response_good"] + d["first_response_bad"]) * 100, 1)
                                   if (d["first_response_good"] + d["first_response_bad"]) > 0 else 0,
            "first_response_avg_h": round(sum(d["first_response_hours"]) / len(d["first_response_hours"]), 1)
                                    if d["first_response_hours"] else 0,
            # Críticos
            "critical_total":     d["critical_total"],
            "critical_sla_good":  d["critical_sla_good"],
            "critical_sla_bad":   d["critical_sla_bad"],
            "critical_sla_pct":   round(d["critical_sla_good"] / (d["critical_sla_good"] + d["critical_sla_bad"]) * 100, 1)
                                  if (d["critical_sla_good"] + d["critical_sla_bad"]) > 0 else 0,
            "critical_mttr":      round(sum(d["critical_res_hours"]) / len(d["critical_res_hours"]), 1)
                                  if d["critical_res_hours"] else 0,
            "critical_by_priority": dict(d["critical_by_priority"]),
            # prev_year_total y total_vs_prev_pct los inyecta el endpoint
            "prev_year_total":    0,
            "total_vs_prev_pct":  None,
            "last30_total":        d["last30_total"],
            "last30_critical":     d["last30_critical"],
            "last30_resolved":     d["last30_resolved"],
            "last30_sla_bad":      d["last30_sla_bad"],
            "last30_sla_good":     d["last30_sla_good"],
            "last30_sla_pct":      round(d["last30_sla_good"] /
                                   (d["last30_sla_good"] + d["last30_sla_bad"]) * 100, 1)
                                   if (d["last30_sla_good"] + d["last30_sla_bad"]) > 0 else 0,
            "last30_sla_breach_pct": round(d["last30_sla_bad"] /
                                   (d["last30_sla_good"] + d["last30_sla_bad"]) * 100, 1)
                                   if (d["last30_sla_good"] + d["last30_sla_bad"]) > 0 else 0,
            # Tiempo Absoluto tickets críticos (promedio en horas)
            "critical_abs_avg_h":  round(sum(d["last30_critical_res_hours"]) /
                                   len(d["last30_critical_res_hours"]), 1)
                                   if d["last30_critical_res_hours"] else 0,
            "critical_abs_p90":    round(sorted(d["last30_critical_res_hours"])[
                                   int(len(d["last30_critical_res_hours"]) * 0.9)], 1)
                                   if d["last30_critical_res_hours"] else 0,
        }
    return result

# ════════════════════════════════════════════════════════
# ENDPOINTS JIRA
# ════════════════════════════════════════════════════════

@app.route("/api/test-connection", methods=["POST"])
def test_connection():
    b = request.json or {}
    url   = b.get("url",   DEFAULT_URL)
    email = b.get("email", DEFAULT_EMAIL)
    token = b.get("token", DEFAULT_TOKEN)
    try:
        me = jira_get(url, email, token, "myself")
        return jsonify({"ok": True, "displayName": me.get("displayName"), "email": me.get("emailAddress")})
    except requests.HTTPError as e:
        return jsonify({"ok": False, "error": f"JIRA {e.response.status_code}: credenciales incorrectas"}), 401
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/projects", methods=["POST"])
def get_projects():
    b = request.json or {}
    url, email, token = b.get("url", DEFAULT_URL), b.get("email", DEFAULT_EMAIL), b.get("token", DEFAULT_TOKEN)
    try:
        data = jira_get(url, email, token, "project/search", {"maxResults": 100, "orderBy": "name"})
        projects = [{"key": p["key"], "name": p["name"]} for p in data.get("values", [])]
        return jsonify({"projects": projects})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fields", methods=["POST"])
def get_fields():
    b = request.json or {}
    url, email, token = b.get("url", DEFAULT_URL), b.get("email", DEFAULT_EMAIL), b.get("token", DEFAULT_TOKEN)
    try:
        fields = jira_get(url, email, token, "field")
        custom = sorted(
            [{"id": f["id"], "name": f["name"]} for f in fields if f["id"].startswith("customfield_")],
            key=lambda x: x["name"].lower()
        )
        return jsonify({"fields": custom})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/debug-sample", methods=["POST"])
def debug_sample():
    """Retorna una muestra de un issue para debuggear campos."""
    b = request.json or {}
    url       = b.get("url",      DEFAULT_URL)
    email     = b.get("email",    DEFAULT_EMAIL)
    token     = b.get("token",    DEFAULT_TOKEN)
    projects  = b.get("projects", [])
    try:
        proj_filter = ("project in (" + ",".join(f'"{p}"' for p in projects) + ")"
                       if projects else "project is not EMPTY")
        jql = f"{proj_filter} ORDER BY created DESC"
        data = jira_get(url, email, token, "search/jql", {
            "jql": jql, "maxResults": 3,
            "fields": "*all"
        })
        issues = data.get("issues", [])
        if not issues:
            return jsonify({"ok": True, "message": "No se encontraron issues", "total": 0})
        # Mostrar campos con datos del primer issue
        f = issues[0]["fields"]
        populated = {k: v for k, v in f.items() if v is not None and v != "" and v != [] and v != {}}
        return jsonify({
            "ok": True,
            "total_in_project": data.get("total", 0),
            "sample_key": issues[0]["key"],
            "sample_status": (f.get("status") or {}).get("name"),
            "populated_fields": list(populated.keys()),
            "sla_candidates": {c: f.get(c) for c in SLA_FIELD_CANDIDATES if f.get(c)},
            "org_candidates": {c: f.get(c) for c in list(ACCOUNT_CANDIDATES) if f.get(c)},
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500

@app.route("/api/metrics", methods=["POST"])
def get_metrics():
    b         = request.json or {}
    url       = b.get("url",       DEFAULT_URL)
    email     = b.get("email",     DEFAULT_EMAIL)
    token     = b.get("token",     DEFAULT_TOKEN)
    projects  = b.get("projects",  [])
    sla_field = b.get("slaField",  "auto")
    days      = b.get("days",      90)
    # Filtro de clientes: lista de nombres separados por coma desde el frontend
    clients_raw = b.get("clients", "")
    client_filter = None
    if clients_raw and clients_raw.strip():
        client_filter = {c.strip() for c in clients_raw.split(",") if c.strip()}
        print(f"[JIRA] Filtro de clientes: {client_filter}")

    if not url or not email or not token:
        return jsonify({"ok": False, "error": "Faltan credenciales"}), 400

    try:
        proj_filter = ("project in (" + ",".join(f'"{p}"' for p in projects) + ")"
                       if projects else "project is not EMPTY")

        fields = list(set([
            "summary", "status", "priority", "issuetype",
            "created", "resolutiondate", "organizations",
            *list(ACCOUNT_CANDIDATES),
            *SLA_FIELD_CANDIDATES,
            *EXTRA_FIELDS,
        ]))
        if sla_field and sla_field.startswith("customfield_") and sla_field not in fields:
            fields.append(sla_field)

        def build_jql(extra=""):
            base = f"{proj_filter}{extra}"
            if client_filter:
                org_names = " OR ".join(f'"{n}"' for n in client_filter)
                return f'({base}) AND (organizationName in ({org_names}) OR cf[11329] in ({org_names}))'
            return base

        # ── Request 1: período actual ────────────────────────
        jql_current = build_jql() + f" AND created >= -{days}d ORDER BY created DESC"
        print(f"\n[JIRA] JQL actual: {jql_current}")
        issues_current = jira_search(url, email, token, jql_current, fields)
        print(f"[JIRA] Issues período actual: {len(issues_current)}")

        # ── Request 2: mismo período del año pasado ──────────
        tz_ar = timezone(timedelta(hours=-3))
        now_ar = datetime.now(tz_ar)
        py_end   = now_ar - timedelta(days=365)
        py_start = py_end  - timedelta(days=days)
        py_start_str = py_start.strftime("%Y-%m-%d")
        py_end_str   = py_end.strftime("%Y-%m-%d")
        jql_prev = build_jql() + f' AND created >= "{py_start_str}" AND created <= "{py_end_str}" ORDER BY created DESC'
        print(f"[JIRA] JQL año anterior: {jql_prev}")
        issues_prev = jira_search(url, email, token, jql_prev, fields)
        print(f"[JIRA] Issues año anterior: {len(issues_prev)}")

        if not issues_current:
            return jsonify({
                "ok": True,
                "total_issues": 0,
                "clients": [],
                "warning": f"No se encontraron issues en los últimos {days} días.",
                "generated_at": datetime.now(tz_ar).isoformat(),
            })

        metrics = process_issues(issues_current, sla_field, client_filter, days)

        # Inyectar conteos del año anterior en cada cliente
        prev_counts = {}
        for issue in issues_prev:
            f_prev = issue["fields"]
            c_prev = "Sin cliente"
            orgs = f_prev.get("organizations")
            if orgs and isinstance(orgs, list) and orgs:
                c_prev = orgs[0].get("name") or orgs[0].get("displayName") or "Sin cliente"
            if c_prev == "Sin cliente":
                cf_org = f_prev.get("customfield_11302")
                if cf_org and isinstance(cf_org, list) and cf_org:
                    c_prev = cf_org[0].get("name") or cf_org[0].get("displayName") or "Sin cliente"
            if client_filter:
                if c_prev not in client_filter and not any(
                    fn.lower() in c_prev.lower() or c_prev.lower() in fn.lower()
                    for fn in client_filter
                ):
                    continue
            prev_counts[c_prev] = prev_counts.get(c_prev, 0) + 1

        # Actualizar métricas con prev_year_total real (mismo período año anterior)
        result_clients = list(metrics.values())
        for client_data in result_clients:
            name = client_data["name"]
            prev = prev_counts.get(name, 0)
            client_data["prev_year_total"] = prev
            client_data["total_vs_prev_pct"] = (
                round((client_data["total"] - prev) / prev * 100, 1) if prev > 0 else None
            )

        response_data = {
            "ok": True,
            "total_issues": len(issues_current),
            "clients": result_clients,
            "generated_at": datetime.now(tz_ar).isoformat(),
            "params": {"projects": projects, "days": days, "clients": clients_raw or ""},
        }
        save_cache(response_data)
        return jsonify(response_data)
    except requests.HTTPError as e:
        body = ""
        try:   body = e.response.json()
        except: body = e.response.text[:500]
        print(f"\n[ERROR JIRA] {e.response.status_code}: {body}\n")
        msg = body.get("errorMessages", [str(body)])[0] if isinstance(body, dict) else str(body)
        # Si el JQL optimizado falla (organizationName puede no existir), reintentar sin filtro de org
        if client_filter and e.response.status_code == 400:
            try:
                jql_fb_cur  = f"{proj_filter} AND created >= -{days}d ORDER BY created DESC"
                print(f"[JIRA] Fallback actual: {jql_fb_cur}")
                issues_current = jira_search(url, email, token, jql_fb_cur, fields)
                tz_ar2 = timezone(timedelta(hours=-3))
                now_ar2 = datetime.now(tz_ar2)
                py_end2   = now_ar2 - timedelta(days=365)
                py_start2 = py_end2  - timedelta(days=days)
                jql_fb_prev = (f'{proj_filter} AND created >= "{py_start2.strftime("%Y-%m-%d")}"'
                               f' AND created <= "{py_end2.strftime("%Y-%m-%d")}" ORDER BY created DESC')
                issues_prev2 = jira_search(url, email, token, jql_fb_prev, fields)
                metrics = process_issues(issues_current, sla_field, client_filter, days)
                prev_counts2 = {}
                for iss in issues_prev2:
                    fp = iss["fields"]
                    cp = "Sin cliente"
                    o2 = fp.get("organizations")
                    if o2 and isinstance(o2, list) and o2:
                        cp = o2[0].get("name") or o2[0].get("displayName") or "Sin cliente"
                    if cp == "Sin cliente":
                        cfo = fp.get("customfield_11302")
                        if cfo and isinstance(cfo, list) and cfo:
                            cp = cfo[0].get("name") or cfo[0].get("displayName") or "Sin cliente"
                    prev_counts2[cp] = prev_counts2.get(cp, 0) + 1
                result2 = list(metrics.values())
                for cd in result2:
                    prev2 = prev_counts2.get(cd["name"], 0)
                    cd["prev_year_total"] = prev2
                    cd["total_vs_prev_pct"] = round((cd["total"] - prev2) / prev2 * 100, 1) if prev2 > 0 else None
                response_data2 = {
                    "ok": True,
                    "total_issues": len(issues_current),
                    "clients": result2,
                    "generated_at": datetime.now(tz_ar2).isoformat(),
                    "warning": "Filtro JQL de organización no soportado — se aplicó filtro local.",
                    "params": {"projects": projects, "days": days, "clients": clients_raw or ""},
                }
                save_cache(response_data2)
                return jsonify(response_data2)
            except Exception as e2:
                return jsonify({"ok": False, "error": str(e2)}), 500
        return jsonify({"ok": False, "error": f"JIRA {e.response.status_code}: {msg}"}), 502
    except Exception as e:
        import traceback
        print(f"\n[ERROR] {traceback.format_exc()}\n")
        return jsonify({"ok": False, "error": str(e)}), 500

# ════════════════════════════════════════════════════════
# CACHE ENDPOINTS
# ════════════════════════════════════════════════════════

@app.route("/api/cache/info", methods=["GET"])
def cache_info():
    c = load_cache()
    if not c:
        return jsonify({"ok": True, "available": False})
    return jsonify({
        "ok": True,
        "available": True,
        "generated_at": c.get("generated_at"),
        "total_issues": c.get("total_issues"),
        "clients_count": len(c.get("clients", [])),
        "projects": c.get("params", {}).get("projects", []),
        "days": c.get("params", {}).get("days"),
        "clients_filter": c.get("params", {}).get("clients", ""),
        "warning": c.get("warning", ""),
    })

@app.route("/api/cache/load", methods=["GET"])
def cache_load():
    c = load_cache()
    if not c:
        return jsonify({"ok": False, "error": "No hay cache disponible"}), 404
    return jsonify({**c, "ok": True, "from_cache": True})

@app.route("/api/cache/clear", methods=["POST"])
def cache_clear():
    try:
        if os.path.exists(CACHE_FILE):
            os.remove(CACHE_FILE)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

# ════════════════════════════════════════════════════════
# EXPORT ENDPOINTS
# ════════════════════════════════════════════════════════

@app.route("/api/export/excel", methods=["POST"])
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl no instalado. Ejecuta: pip install openpyxl"}), 500

    b = request.json or {}
    clients = b.get("clients", [])

    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen ─────────────────────────────────
    ws = wb.active
    ws.title = "Resumen Clientes"
    navy = "0F2044"
    header_fill = PatternFill("solid", fgColor=navy)
    header_font = Font(color="FFFFFF", bold=True, size=11)
    headers = ["Cliente", "Health", "SLA %", "MTTR (h)", "Abiertos",
               ">7d", ">14d", "Bug Rate %", "Total", "Críticos SLA", "Reapertura %"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, c in enumerate(clients, 2):
        ws.cell(row=row, column=1, value=c.get("name"))
        ws.cell(row=row, column=2, value=c.get("health"))
        ws.cell(row=row, column=3, value=c.get("sla_pct"))
        ws.cell(row=row, column=4, value=c.get("mttr"))
        ws.cell(row=row, column=5, value=c.get("open"))
        ws.cell(row=row, column=6, value=c.get("aging_7"))
        ws.cell(row=row, column=7, value=c.get("aging_14"))
        ws.cell(row=row, column=8, value=c.get("bug_rate"))
        ws.cell(row=row, column=9, value=c.get("total"))
        ws.cell(row=row, column=10, value=c.get("sla_bad"))
        ws.cell(row=row, column=11, value=c.get("reopen_rate"))

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # ── Hoja 2: Por Trimestre ───────────────────────────
    ws2 = wb.create_sheet("Por Trimestre")
    ws2.cell(row=1, column=1, value="Cliente").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2, value="Trimestre").font = header_font
    ws2.cell(row=1, column=2).fill = header_fill
    ws2.cell(row=1, column=3, value="Tickets").font = header_font
    ws2.cell(row=1, column=3).fill = header_fill
    row2 = 2
    for c in clients:
        for q, v in (c.get("by_quarter") or {}).items():
            ws2.cell(row=row2, column=1, value=c["name"])
            ws2.cell(row=row2, column=2, value=q)
            ws2.cell(row=row2, column=3, value=v)
            row2 += 1
    for col in [1, 2, 3]:
        ws2.column_dimensions[get_column_letter(col)].width = 22

    # ── Hoja 3: Por Tipo ────────────────────────────────
    ws3 = wb.create_sheet("Por Tipo y Prioridad")
    ws3.cell(row=1, column=1, value="Cliente").font = header_font
    ws3.cell(row=1, column=1).fill = header_fill
    ws3.cell(row=1, column=2, value="Tipo").font = header_font
    ws3.cell(row=1, column=2).fill = header_fill
    ws3.cell(row=1, column=3, value="Cantidad").font = header_font
    ws3.cell(row=1, column=3).fill = header_fill
    row3 = 2
    for c in clients:
        for tipo, v in (c.get("by_type") or {}).items():
            ws3.cell(row=row3, column=1, value=c["name"])
            ws3.cell(row=row3, column=2, value=tipo)
            ws3.cell(row=row3, column=3, value=v)
            row3 += 1
    for col in [1, 2, 3]:
        ws3.column_dimensions[get_column_letter(col)].width = 22

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"csm_dashboard_{datetime.now().strftime('%Y%m%d')}.xlsx")

@app.route("/api/export/pdf", methods=["POST"])
def export_pdf():
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_CENTER
    except ImportError:
        return jsonify({"error": "reportlab no instalado. Ejecuta: pip install reportlab"}), 500

    b = request.json or {}
    clients = b.get("clients", [])

    out = io.BytesIO()
    doc = SimpleDocTemplate(out, pagesize=landscape(A4),
                            rightMargin=1*cm, leftMargin=1*cm, topMargin=1.5*cm, bottomMargin=1*cm)

    styles = getSampleStyleSheet()
    navy = colors.HexColor("#0F2044")
    elements = []

    title_style = ParagraphStyle("title", parent=styles["Heading1"],
                                 textColor=navy, fontSize=18, spaceAfter=8)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"],
                               textColor=colors.HexColor("#64748b"), fontSize=9, spaceAfter=20)

    elements.append(Paragraph("CSM Dashboard — Reporte de Clientes", title_style))
    elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} · {len(clients)} clientes", sub_style))

    col_widths = [4.5*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm, 2*cm, 1.8*cm]
    data = [["Cliente", "Health", "SLA %", "MTTR", "Abiertos", ">7d", ">14d", "Bug Rate", "Total"]]
    for c in clients:
        data.append([
            c.get("name","")[:30],
            str(c.get("health", 0)),
            f"{c.get('sla_pct',0)}%",
            f"{c.get('mttr',0)}h",
            str(c.get("open", 0)),
            str(c.get("aging_7", 0)),
            str(c.get("aging_14", 0)),
            f"{c.get('bug_rate',0)}%",
            str(c.get("total", 0)),
        ])

    t = Table(data, colWidths=col_widths)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), navy),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,-1), 9),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f8fafc")]),
        ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#e2e8f0")),
        ("ALIGN", (1,0), (-1,-1), "CENTER"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
    ]))

    # Colorear health score
    for i, c in enumerate(clients, 1):
        h = c.get("health", 0)
        color = colors.HexColor("#10b981") if h >= 70 else (colors.HexColor("#f59e0b") if h >= 50 else colors.HexColor("#ef4444"))
        t.setStyle(TableStyle([("TEXTCOLOR", (1,i), (1,i), color), ("FONTNAME", (1,i), (1,i), "Helvetica-Bold")]))

    elements.append(t)
    doc.build(elements)
    out.seek(0)
    return send_file(out, mimetype="application/pdf",
                     as_attachment=True, download_name=f"csm_dashboard_{datetime.now().strftime('%Y%m%d')}.pdf")

# ════════════════════════════════════════════════════════
# MYSQL ENDPOINT (opcional)
# ════════════════════════════════════════════════════════

@app.route("/api/mysql/query", methods=["POST"])
def mysql_query():
    try:
        import mysql.connector
    except ImportError:
        return jsonify({"error": "mysql-connector-python no instalado. Ejecuta: pip install mysql-connector-python"}), 500

    b = request.json or {}
    host  = b.get("host",     "localhost")
    port  = b.get("port",     3306)
    user  = b.get("user",     "")
    pwd   = b.get("password", "")
    db    = b.get("database", "")
    query = b.get("query",    "")

    try:
        conn = mysql.connector.connect(host=host, port=port, user=user, password=pwd, database=db, connection_timeout=10)
        cursor = conn.cursor(dictionary=True)
        cursor.execute(query)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify({"ok": True, "rows": rows, "count": len(rows)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/mysql/test", methods=["POST"])
def mysql_test():
    try:
        import mysql.connector
    except ImportError:
        return jsonify({"ok": False, "error": "mysql-connector-python no instalado"}), 500
    b = request.json or {}
    try:
        conn = mysql.connector.connect(
            host=b.get("host", "localhost"), port=b.get("port", 3306),
            user=b.get("user", ""), password=b.get("password", ""),
            database=b.get("database", ""), connection_timeout=10
        )
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

# ════════════════════════════════════════════════════════
if __name__ == "__main__":
    try:
        ip = socket.gethostbyname(socket.gethostname())
    except Exception:
        ip = "TU_IP_LOCAL"
    print("=" * 55)
    print("  CSM Dashboard v2 — Listo")
    print("=" * 55)
    print(f"\n  Tu máquina : http://localhost:5050")
    print(f"  Tu equipo  : http://{ip}:5050")
    print("\n  Instalar dependencias opcionales:")
    print("    pip install openpyxl reportlab mysql-connector-python")
    print()
    app.run(host="0.0.0.0", port=5050, debug=False)
